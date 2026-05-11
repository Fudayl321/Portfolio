"""
Bank Marketing Dashboard — Rebuilt with Real Excel Formulas
============================================================
Every KPI and summary table uses actual Excel formulas (COUNTIFS, SUMIFS,
AVERAGEIFS, IFERROR, named ranges) pulling from the raw data table.
No hardcoded values. Anyone opening the file can see exactly how each
number is calculated.

Author : Fudayl Abdul Jalil
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
import os, warnings
warnings.filterwarnings('ignore')

df = pd.read_csv("/home/claude/portfolio/02_Bank_Marketing_Analysis/02_cleaning/bank_marketing_clean.csv")
OUT = "/home/claude/portfolio/02_Bank_Marketing_Analysis/05_output/bank_marketing_dashboard.xlsx"

# ── STYLE CONSTANTS ───────────────────────────────────────────
NAVY   = "1A3C5E"; GREEN  = "217346"; RED    = "C0392B"
AMBER  = "E67E22"; GREY   = "595959"; LGREY  = "F2F2F2"
WHITE  = "FFFFFF"; LBLUE  = "D6E4F0"
thin   = Side(style='thin', color='CCCCCC')
bdr    = Border(top=thin, left=thin, right=thin, bottom=thin)

def hdr(ws, r, c, txt, bg=NAVY, fg=WHITE, sz=10, bold=True, align='center', wrap=False, colspan=1):
    if colspan > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+colspan-1)
    cell = ws.cell(r, c, txt)
    cell.font      = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = bdr
    return cell

def cel(ws, r, c, val=None, formula=None, fmt=None, bold=False,
        fg='000000', align='center', bg=None, sz=10):
    cell = ws.cell(r, c, formula if formula else val)
    cell.font      = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = bdr
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def section(ws, r, c, txt, span=8):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    cell = ws.cell(r, c, txt)
    cell.font      = Font(name='Calibri', bold=True, color=NAVY, size=12)
    cell.fill      = PatternFill("solid", fgColor=LBLUE)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22

def kpi_block(ws, r, c, label, formula, fmt='#,##0', color=GREEN):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
    lc = ws.cell(r, c, label)
    lc.font = Font(name='Calibri', color=GREY, size=9)
    lc.fill = PatternFill("solid", fgColor=LGREY)
    lc.alignment = Alignment(horizontal='center', vertical='center')
    vc = ws.cell(r+1, c, formula)
    vc.font = Font(name='Calibri', bold=True, color=color, size=20)
    vc.number_format = fmt
    vc.fill = PatternFill("solid", fgColor=LGREY)
    vc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height   = 16
    ws.row_dimensions[r+1].height = 34

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — RAW DATA  (named table so formulas reference it)
# ══════════════════════════════════════════════════════════════
ws_raw = wb.active
ws_raw.title = "Raw Data"
ws_raw.sheet_properties.tabColor = "808080"
ws_raw.sheet_view.showGridLines = True

# Write only the columns used in the dashboard formulas
cols = ['age','job','marital','education','contact','month',
        'duration_min','campaign','poutcome','subscribed',
        'age_group','contacted_before','y']

for ci, col in enumerate(cols, 1):
    hdr(ws_raw, 1, ci, col.replace('_',' ').title(), bg=NAVY, sz=9)

for ri, row in enumerate(df[cols].itertuples(index=False), 2):
    alt = LGREY if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        c = ws_raw.cell(ri, ci, val)
        c.font      = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='center')
        c.fill      = PatternFill("solid", fgColor=alt)

# Auto-filter on raw data
last_col = get_column_letter(len(cols))
ws_raw.auto_filter.ref = f"A1:{last_col}1"

# Set column widths
for ci, w in enumerate([8,16,12,18,12,8,14,10,14,12,12,16,8], 1):
    ws_raw.column_dimensions[get_column_letter(ci)].width = w

# Named ranges for formula use — reference the raw data columns
# Job=col B, Subscribed=col J (10), Age_group=col K (11), Poutcome=col I (9)
# Month=col F (6), Contact=col E (5), Duration_min=col G (7)
total_rows = len(df) + 1   # +1 for header

RAW = "Raw Data"   # sheet name for cross-sheet formula references
J   = "J"          # subscribed column
B   = "B"          # job
F   = "F"          # month
K   = "K"          # age_group
I_  = "I"          # poutcome
E   = "E"          # contact
G   = "G"          # duration_min
C   = "C"          # marital
D   = "D"          # education
N   = str(total_rows)  # last data row

def ref(col): return f"'{RAW}'!{col}2:{col}{N}"

# ══════════════════════════════════════════════════════════════
# SHEET 2 — LOOKUP TABLES  (what the dashboard formulas draw from)
# ══════════════════════════════════════════════════════════════
ws_lk = wb.create_sheet("Lookup Tables")
ws_lk.sheet_properties.tabColor = AMBER
ws_lk.sheet_view.showGridLines = False

hdr(ws_lk, 1, 1, "LOOKUP TABLES — Source for all Dashboard formulas", bg=NAVY, sz=11, colspan=12)
ws_lk.row_dimensions[1].height = 28
ws_lk.merge_cells("A1:L1")

# ── Table 1: Job lookup (cols A-E) ───────────────────────────
section(ws_lk, 3, 1, "  JOB SUBSCRIPTION ANALYSIS", 5)
for ci, h in enumerate(['Job','Total','Subscriptions','Sub Rate %','Avg Call (min)'], 1):
    hdr(ws_lk, 4, ci, h, bg=GREEN)

jobs = sorted(df['job'].unique())
for ri, job in enumerate(jobs, 5):
    r = ri
    jq = f'"{job}"'
    ws_lk.cell(r, 1, job).font = Font(name='Calibri', size=10, bold=True)
    ws_lk.cell(r, 1).alignment = Alignment(horizontal='left')
    ws_lk.cell(r, 1).border = bdr

    # COUNTIFS — count all rows where job matches
    cel(ws_lk, r, 2,
        formula=f"=COUNTIFS('{RAW}'!{B}2:{B}{N},{jq})",
        fmt='#,##0')

    # COUNTIFS — count rows where job matches AND subscribed=1
    cel(ws_lk, r, 3,
        formula=f"=COUNTIFS('{RAW}'!{B}2:{B}{N},{jq},'{RAW}'!{J}2:{J}{N},1)",
        fmt='#,##0')

    # IFERROR(subscriptions/total, 0) — sub rate
    cel(ws_lk, r, 4,
        formula=f"=IFERROR(C{r}/B{r},0)",
        fmt='0.0%')

    # AVERAGEIFS — avg duration where job matches
    cel(ws_lk, r, 5,
        formula=f"=IFERROR(AVERAGEIFS('{RAW}'!{G}2:{G}{N},'{RAW}'!{B}2:{B}{N},{jq}),0)",
        fmt='0.0')

    for ci in range(1, 6):
        ws_lk.cell(r, ci).border = bdr
        if ci > 1:
            ws_lk.cell(r, ci).alignment = Alignment(horizontal='center')

# ── Table 2: Month lookup (cols G-K) ─────────────────────────
section(ws_lk, 3, 7, "  MONTHLY PERFORMANCE", 5)
for ci, h in enumerate(['Month','Month Num','Total','Subscriptions','Sub Rate %'], 7):
    hdr(ws_lk, 4, ci, h, bg=NAVY)

month_order = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
for ri, (mo, mn) in enumerate(zip(month_order, range(1,13)), 5):
    mq = f'"{mo}"'
    ws_lk.cell(ri, 7, mo.title()).font = Font(name='Calibri', size=10, bold=True)
    ws_lk.cell(ri, 7).alignment = Alignment(horizontal='left')
    ws_lk.cell(ri, 7).border = bdr
    cel(ws_lk, ri, 8,  val=mn, fmt='0')
    cel(ws_lk, ri, 9,  formula=f"=COUNTIFS('{RAW}'!{F}2:{F}{N},{mq})", fmt='#,##0')
    cel(ws_lk, ri, 10, formula=f"=COUNTIFS('{RAW}'!{F}2:{F}{N},{mq},'{RAW}'!{J}2:{J}{N},1)", fmt='#,##0')
    cel(ws_lk, ri, 11, formula=f"=IFERROR(J{ri}/I{ri},0)", fmt='0.0%')
    for ci in range(7, 12):
        ws_lk.cell(ri, ci).border = bdr
        if ci > 7:
            ws_lk.cell(ri, ci).alignment = Alignment(horizontal='center')

# ── Table 3: Age group lookup (cols A-D, below job table) ────
r_age = 5 + len(jobs) + 2
section(ws_lk, r_age, 1, "  AGE GROUP ANALYSIS", 5)
for ci, h in enumerate(['Age Group','Total','Subscriptions','Sub Rate %','Avg Call (min)'], 1):
    hdr(ws_lk, r_age+1, ci, h, bg=GREEN)

age_groups = ['<25','25-34','35-44','45-54','55-64','65+']
for ri, ag in enumerate(age_groups, r_age+2):
    aq = f'"{ag}"'
    ws_lk.cell(ri, 1, ag).font = Font(name='Calibri', size=10, bold=True)
    ws_lk.cell(ri, 1).alignment = Alignment(horizontal='left')
    ws_lk.cell(ri, 1).border = bdr
    cel(ws_lk, ri, 2, formula=f"=COUNTIFS('{RAW}'!{K}2:{K}{N},{aq})", fmt='#,##0')
    cel(ws_lk, ri, 3, formula=f"=COUNTIFS('{RAW}'!{K}2:{K}{N},{aq},'{RAW}'!{J}2:{J}{N},1)", fmt='#,##0')
    cel(ws_lk, ri, 4, formula=f"=IFERROR(C{ri}/B{ri},0)", fmt='0.0%')
    cel(ws_lk, ri, 5,
        formula=f"=IFERROR(AVERAGEIFS('{RAW}'!{G}2:{G}{N},'{RAW}'!{K}2:{K}{N},{aq}),0)",
        fmt='0.0')
    for ci in range(1, 6):
        ws_lk.cell(ri, ci).border = bdr
        if ci > 1:
            ws_lk.cell(ri, ci).alignment = Alignment(horizontal='center')

# ── Table 4: Previous outcome (cols G-J, below month table) ──
r_po = 17 + 2
section(ws_lk, r_po, 7, "  PREVIOUS OUTCOME IMPACT", 5)
for ci, h in enumerate(['Prev Outcome','Total','Subscriptions','Sub Rate %'], 7):
    hdr(ws_lk, r_po+1, ci, h, bg=NAVY)

outcomes = ['failure','nonexistent','success']
for ri, po in enumerate(outcomes, r_po+2):
    pq = f'"{po}"'
    ws_lk.cell(ri, 7, po.title()).font = Font(name='Calibri', size=10, bold=True)
    ws_lk.cell(ri, 7).alignment = Alignment(horizontal='left')
    ws_lk.cell(ri, 7).border = bdr
    cel(ws_lk, ri, 8,  formula=f"=COUNTIFS('{RAW}'!{I_}2:{I_}{N},{pq})", fmt='#,##0')
    cel(ws_lk, ri, 9,  formula=f"=COUNTIFS('{RAW}'!{I_}2:{I_}{N},{pq},'{RAW}'!{J}2:{J}{N},1)", fmt='#,##0')
    cel(ws_lk, ri, 10, formula=f"=IFERROR(I{ri}/H{ri},0)", fmt='0.0%')
    for ci in range(7, 11):
        ws_lk.cell(ri, ci).border = bdr
        if ci > 7:
            ws_lk.cell(ri, ci).alignment = Alignment(horizontal='center')

for col_l, w in zip('ABCDEFGHIJK', [18,12,14,10,14,4,12,10,12,12,12]):
    ws_lk.column_dimensions[col_l].width = w

# ══════════════════════════════════════════════════════════════
# SHEET 3 — EXECUTIVE DASHBOARD  (all cells reference Lookup Tables)
# ══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Executive Dashboard")
ws1.sheet_properties.tabColor = NAVY
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:M3")
t = ws1["A1"]
t.value = "BANK MARKETING CAMPAIGN ANALYTICS  |  Subscription Conversion & Customer Insights"
t.font  = Font(name='Calibri', bold=True, color=WHITE, size=15)
t.fill  = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 38
ws1.row_dimensions[2].height = 16
ws1.row_dimensions[3].height = 10

# Source note
ws1.merge_cells("A4:M4")
src = ws1["A4"]
src.value = "Data source: UCI ML Repository — Bank Marketing  |  Formulas pull live from 'Raw Data' sheet  |  Author: Fudayl Abdul Jalil"
src.font  = Font(name='Calibri', color=NAVY, size=9, italic=True)
src.fill  = PatternFill("solid", fgColor=LBLUE)
src.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[4].height = 14

# ── KPI cards using COUNTIF/COUNTIFS formulas ─────────────────
ws1.row_dimensions[5].height = 8

# Total contacts — COUNTA on job column (excludes header)
kpi_block(ws1, 6, 1,  "Total Contacts",
          f"=COUNTA('{RAW}'!{B}2:{B}{N})", '#,##0', NAVY)

# Subscriptions — COUNTIF where subscribed=1
kpi_block(ws1, 6, 3,  "Subscriptions",
          f"=COUNTIF('{RAW}'!{J}2:{J}{N},1)", '#,##0', GREEN)

# Conversion rate — subscriptions / total
kpi_block(ws1, 6, 5,  "Conversion Rate",
          f"=IFERROR(C7/A7,0)", '0.0%', RED)

# Avg age — AVERAGE
kpi_block(ws1, 6, 7,  "Avg Client Age",
          f"=AVERAGE('{RAW}'!A2:A{N})", '0.0 "yrs"', NAVY)

# Avg call duration — AVERAGE
kpi_block(ws1, 6, 9,  "Avg Call Duration",
          f"=AVERAGE('{RAW}'!{G}2:{G}{N})", '0.0 "min"', GREEN)

# Avg campaign contacts
kpi_block(ws1, 6, 11, "Avg Campaign Contacts",
          f"=AVERAGE('{RAW}'!H2:H{N})", '0.0"x"', AMBER)

ws1.row_dimensions[8].height = 14

# ── JOB TABLE — formulas pull from Lookup Tables ──────────────
section(ws1, 9, 1, "  SUBSCRIPTION RATE BY JOB TYPE  (formulas → Lookup Tables sheet)", 10)
hdrs = ['Job','Total Contacts','Subscriptions','Sub Rate %','Avg Call (min)','Rank']
for ci, h in enumerate(hdrs, 1):
    hdr(ws1, 10, ci, h, bg=GREEN)

# Pull job data from Lookup Tables using INDEX/MATCH
# Sort is handled by referencing the Lookup Tables rows directly
# (already sorted by job name in the lookup table)
for ri, job in enumerate(jobs, 11):
    lk_row = 4 + jobs.index(job) + 1  # row in Lookup Tables sheet
    alt = LGREY if ri % 2 == 0 else WHITE

    # Job name — direct reference
    cel(ws1, ri, 1,
        formula=f"='Lookup Tables'!A{lk_row}",
        bold=True, align='left', bg=alt)

    # Total — reference lookup
    cel(ws1, ri, 2,
        formula=f"='Lookup Tables'!B{lk_row}",
        fmt='#,##0', bg=alt)

    # Subscriptions
    cel(ws1, ri, 3,
        formula=f"='Lookup Tables'!C{lk_row}",
        fmt='#,##0', bg=alt)

    # Sub rate — reference lookup (already has formula there)
    cel(ws1, ri, 4,
        formula=f"='Lookup Tables'!D{lk_row}",
        fmt='0.0%', bg=alt)

    # Avg call
    cel(ws1, ri, 5,
        formula=f"='Lookup Tables'!E{lk_row}",
        fmt='0.0', bg=alt)

    # RANK formula — shows position of this job's sub rate vs all others
    job_sub_range = f"'Lookup Tables'!D5:D{4+len(jobs)}"
    cel(ws1, ri, 6,
        formula=f"=RANK(D{ri},{job_sub_range},0)",
        fmt='0', bold=True, bg=alt)

# Conditional formatting on sub rate column (D11:D{10+len(jobs)})
sub_rate_range = f"D11:D{10+len(jobs)}"
ws1.conditional_formatting.add(sub_rate_range, ColorScaleRule(
    start_type='min', start_value=0, start_color='FF6B6B',
    mid_type='percentile', mid_value=50, mid_color='FFD93D',
    end_type='max', end_value=0, end_color='6BCB77'))

job_last_row = 10 + len(jobs)

# ── MONTHLY TABLE ─────────────────────────────────────────────
r0 = job_last_row + 2
section(ws1, r0, 1, "  MONTHLY CAMPAIGN PERFORMANCE  (formulas → Lookup Tables sheet)", 10)
for ci, h in enumerate(['Month','Total Contacts','Subscriptions','Sub Rate %','vs Avg'], 1):
    hdr(ws1, r0+1, ci, h, bg=NAVY)

for ri, (mo, lk_row) in enumerate(zip(month_order, range(5, 17)), r0+2):
    alt = LGREY if ri % 2 == 0 else WHITE
    cel(ws1, ri, 1, formula=f"=PROPER('Lookup Tables'!G{lk_row})", align='left', bg=alt)
    cel(ws1, ri, 2, formula=f"='Lookup Tables'!I{lk_row}", fmt='#,##0', bg=alt)
    cel(ws1, ri, 3, formula=f"='Lookup Tables'!J{lk_row}", fmt='#,##0', bg=alt)
    cel(ws1, ri, 4, formula=f"='Lookup Tables'!K{lk_row}", fmt='0.0%', bg=alt)
    # vs average — shows how each month compares to overall avg
    cel(ws1, ri, 5,
        formula=f"=D{ri}-$E$7",  # E7 = overall conversion rate KPI
        fmt='+0.0%;-0.0%',
        fg=GREEN if df[df['month']==month_order[ri-r0-2]]['subscribed'].mean() > df['subscribed'].mean() else RED,
        bg=alt)

month_last_row = r0 + 1 + 12

# ── CHARTS ────────────────────────────────────────────────────
# Bar chart — job sub rate, pulling from Lookup Tables
chart1 = BarChart()
chart1.type = "bar"; chart1.title = "Subscription Rate by Job Type"
chart1.y_axis.title = "Sub Rate %"; chart1.width = 18; chart1.height = 14; chart1.style = 10
rd1 = Reference(ws_lk, min_col=4, max_col=4, min_row=4, max_row=4+len(jobs))
rc1 = Reference(ws_lk, min_col=1, min_row=5, max_row=4+len(jobs))
chart1.add_data(rd1, titles_from_data=True)
chart1.set_categories(rc1)
chart1.series[0].graphicalProperties.solidFill = NAVY
ws1.add_chart(chart1, "H9")

# Line chart — monthly sub rate
chart2 = LineChart()
chart2.title = "Monthly Subscription Rate"; chart2.width = 18; chart2.height = 11; chart2.style = 10
rd2 = Reference(ws_lk, min_col=11, max_col=11, min_row=4, max_row=16)
rc2 = Reference(ws_lk, min_col=7, min_row=5, max_row=16)
chart2.add_data(rd2, titles_from_data=True)
chart2.set_categories(rc2)
chart2.series[0].graphicalProperties.line.solidFill = GREEN
chart2.series[0].graphicalProperties.line.width = 25000
ws1.add_chart(chart2, f"H{r0}")

for col_l, w in zip('ABCDEFGHIJKLM', [18,14,14,10,12,8,4,12,14,14,14,14,6]):
    ws1.column_dimensions[col_l].width = w

# ══════════════════════════════════════════════════════════════
# SHEET 4 — SEGMENT ANALYSIS  (formulas reference Lookup Tables)
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Segment Analysis")
ws2.sheet_properties.tabColor = GREEN
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:J2")
t2 = ws2["A1"]
t2.value = "CUSTOMER SEGMENT ANALYSIS — Age Group, Education & Previous Campaign"
t2.font  = Font(name='Calibri', bold=True, color=WHITE, size=13)
t2.fill  = PatternFill("solid", fgColor=GREEN)
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

# Age group table — pulls from Lookup Tables
r_ag = r_age + 1   # row in Lookup Tables where age table header is
section(ws2, 3, 1, "  AGE GROUP — Sub Rate & Avg Call Duration  (AVERAGEIFS/COUNTIFS)", 8)
for ci, h in enumerate(['Age Group','Total','Subscriptions','Sub Rate %','Avg Call (min)',
                         'vs Overall Avg'], 1):
    hdr(ws2, 4, ci, h, bg=NAVY)

for ri, (ag, lk_row) in enumerate(zip(age_groups, range(r_age+2, r_age+2+len(age_groups))), 5):
    alt = LGREY if ri % 2 == 0 else WHITE
    cel(ws2, ri, 1, formula=f"='Lookup Tables'!A{lk_row}", bold=True, align='left', bg=alt)
    cel(ws2, ri, 2, formula=f"='Lookup Tables'!B{lk_row}", fmt='#,##0', bg=alt)
    cel(ws2, ri, 3, formula=f"='Lookup Tables'!C{lk_row}", fmt='#,##0', bg=alt)
    cel(ws2, ri, 4, formula=f"='Lookup Tables'!D{lk_row}", fmt='0.0%', bg=alt)
    cel(ws2, ri, 5, formula=f"='Lookup Tables'!E{lk_row}", fmt='0.0', bg=alt)
    cel(ws2, ri, 6,
        formula=f"=D{ri}-'Executive Dashboard'!E7",
        fmt='+0.0%;-0.0%', bg=alt)

ws2.conditional_formatting.add(
    f"D5:D{4+len(age_groups)}",
    DataBarRule(start_type='min', start_value=0, end_type='max', end_value=0, color=GREEN))

# Previous outcome table
r_out = 5 + len(age_groups) + 2
section(ws2, r_out, 1, "  PREVIOUS CAMPAIGN OUTCOME  (COUNTIFS)", 8)
for ci, h in enumerate(['Previous Outcome','Total','Subscriptions','Sub Rate %','Multiplier vs Overall'], 1):
    hdr(ws2, r_out+1, ci, h, bg=NAVY)

for ri, (po, lk_row) in enumerate(zip(outcomes, range(r_po+2, r_po+2+len(outcomes))), r_out+2):
    alt = LGREY if ri % 2 == 0 else WHITE
    cel(ws2, ri, 1, formula=f"='Lookup Tables'!G{lk_row}", bold=True, align='left', bg=alt)
    cel(ws2, ri, 2, formula=f"='Lookup Tables'!H{lk_row}", fmt='#,##0', bg=alt)
    cel(ws2, ri, 3, formula=f"='Lookup Tables'!I{lk_row}", fmt='#,##0', bg=alt)
    cel(ws2, ri, 4, formula=f"='Lookup Tables'!J{lk_row}", fmt='0.0%', bg=alt)
    # Multiplier — how many times better than overall avg
    cel(ws2, ri, 5,
        formula=f"=IFERROR(D{ri}/'Executive Dashboard'!E7,0)",
        fmt='0.0"x"', bold=True, bg=alt)

# Education analysis — live COUNTIFS directly on Raw Data
r_edu = r_out + len(outcomes) + 3
section(ws2, r_edu, 1, "  EDUCATION LEVEL ANALYSIS  (direct COUNTIFS on Raw Data)", 8)
for ci, h in enumerate(['Education','Total','Subscriptions','Sub Rate %','Avg Age'], 1):
    hdr(ws2, r_edu+1, ci, h, bg=GREEN)

educations = sorted(df['education'].unique())
for ri, edu in enumerate(educations, r_edu+2):
    eq = f'"{edu}"'
    alt = LGREY if ri % 2 == 0 else WHITE
    ws2.cell(ri, 1, edu.replace('.',' ').title()).font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(ri, 1).alignment = Alignment(horizontal='left')
    ws2.cell(ri, 1).border = bdr
    cel(ws2, ri, 2,
        formula=f"=COUNTIF('{RAW}'!{D}2:{D}{N},{eq})",
        fmt='#,##0', bg=alt)
    cel(ws2, ri, 3,
        formula=f"=COUNTIFS('{RAW}'!{D}2:{D}{N},{eq},'{RAW}'!{J}2:{J}{N},1)",
        fmt='#,##0', bg=alt)
    cel(ws2, ri, 4,
        formula=f"=IFERROR(C{ri}/B{ri},0)",
        fmt='0.0%', bg=alt)
    cel(ws2, ri, 5,
        formula=f"=IFERROR(AVERAGEIFS('{RAW}'!A2:A{N},'{RAW}'!{D}2:{D}{N},{eq}),0)",
        fmt='0.0', bg=alt)
    for ci in range(1, 6):
        ws2.cell(ri, ci).border = bdr
        if ci > 1:
            ws2.cell(ri, ci).alignment = Alignment(horizontal='center')

for col_l, w in zip('ABCDEFGHIJ', [22,12,14,10,14,14,4,14,14,10]):
    ws2.column_dimensions[col_l].width = w

# ══════════════════════════════════════════════════════════════
# SHEET 5 — FORMULA GUIDE  (explains every formula used)
# ══════════════════════════════════════════════════════════════
ws_fg = wb.create_sheet("Formula Guide")
ws_fg.sheet_properties.tabColor = "4472C4"
ws_fg.sheet_view.showGridLines = False

ws_fg.merge_cells("A1:F2")
t_fg = ws_fg["A1"]
t_fg.value = "FORMULA GUIDE — Every formula used in this workbook, explained"
t_fg.font  = Font(name='Calibri', bold=True, color=WHITE, size=13)
t_fg.fill  = PatternFill("solid", fgColor="4472C4")
t_fg.alignment = Alignment(horizontal='center', vertical='center')
ws_fg.row_dimensions[1].height = 32

formulas = [
    ("Formula", "Where Used", "What it Does", "Example"),
    ("COUNTA(range)", "KPI — Total Contacts",
     "Counts all non-empty cells in a range. Used instead of COUNT because COUNT only works on numbers.",
     f"=COUNTA('{RAW}'!B2:B{N})"),
    ("COUNTIF(range, criteria)", "KPI — Subscriptions",
     "Counts cells that match a single condition.",
     f"=COUNTIF('{RAW}'!J2:J{N},1)"),
    ("COUNTIFS(range1, criteria1, range2, criteria2)", "Lookup Tables — Job/Month analysis",
     "Counts rows that match multiple conditions at once. E.g. job='admin' AND subscribed=1.",
     f"=COUNTIFS('{RAW}'!B2:B{N},\"admin.\",'{RAW}'!J2:J{N},1)"),
    ("AVERAGEIFS(avg_range, range1, criteria1, ...)", "Lookup Tables — Avg call by job",
     "Calculates average of a range where multiple conditions are met.",
     f"=AVERAGEIFS('{RAW}'!G2:G{N},'{RAW}'!B2:B{N},\"admin.\")"),
    ("IFERROR(formula, value_if_error)", "All division formulas",
     "Returns a fallback value if the formula errors. Used on all divisions to handle zero denominators.",
     "=IFERROR(C11/B11,0)"),
    ("RANK(value, ref, order)", "Job table — Rank column",
     "Returns the rank of a value within a range. Order=0 means highest=rank 1.",
     "=RANK(D11,'Lookup Tables'!D5:D16,0)"),
    ("PROPER(text)", "Month names on dashboard",
     "Capitalises the first letter of each word. jan → Jan.",
     "=PROPER('Lookup Tables'!G5)"),
    ("Sheet reference: 'Sheet Name'!CellRef", "Dashboard pulling from Lookup Tables",
     "References a cell on another sheet. Single quotes needed when sheet name has spaces.",
     "='Lookup Tables'!B5"),
    ("Cross-sheet formula: =D11-'Executive Dashboard'!E7", "vs Avg columns",
     "Subtracts the overall conversion rate KPI from each row's rate to show above/below average.",
     "=D5-'Executive Dashboard'!E7"),
    ("Conditional formatting — Color Scale", "Sub Rate % columns",
     "Automatically colours cells from red (low) to yellow (mid) to green (high) based on value. Applied via Home → Conditional Formatting → Color Scales.",
     "Applied to D11:D22 on Executive Dashboard"),
    ("Conditional formatting — Data Bars", "Age group sub rate",
     "Fills cells with a bar proportional to their value — works like an in-cell bar chart.",
     "Applied to D5:D10 on Segment Analysis"),
]

for ci, h in enumerate(['Formula','Where Used','What it Does','Example'], 1):
    hdr(ws_fg, 3, ci, h, bg=NAVY, sz=10)

for ri, row in enumerate(formulas[1:], 4):
    alt = LGREY if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        c = ws_fg.cell(ri, ci, val)
        c.font = Font(size=9,
                      name='Courier New' if ci in [1,4] else 'Calibri')
        if ci in [1, 4]:
            c.font = Font(name='Courier New', size=9)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.fill = PatternFill("solid", fgColor=alt)
        c.border = bdr
    ws_fg.row_dimensions[ri].height = 40

for col_l3, w in zip('ABCD', [32, 22, 40, 50]):
    ws_fg.column_dimensions[col_l3].width = w

# ── Sheet order ───────────────────────────────────────────────
wb.move_sheet("Executive Dashboard", offset=-3)
wb.move_sheet("Segment Analysis",    offset=-1)

wb.save(OUT)
print(f"✓ Excel saved → {OUT}")
sizes = {}
wb2 = openpyxl.load_workbook(OUT)
for ws in wb2.worksheets:
    print(f"  Sheet: {ws.title:30s} rows={ws.max_row}")

# Spot check — verify a formula cell
ws_check = wb2["Executive Dashboard"]
print(f"\n  KPI cell A7 (Total Contacts formula): {ws_check['A7'].value}")
print(f"  KPI cell C7 (Subscriptions formula):  {ws_check['C7'].value}")
print(f"  KPI cell E7 (Conversion Rate formula): {ws_check['E7'].value}")
print(f"  Job sub rate D11 formula:             {ws_check['D11'].value}")
print(f"  Rank formula F11:                     {ws_check['F11'].value}")
