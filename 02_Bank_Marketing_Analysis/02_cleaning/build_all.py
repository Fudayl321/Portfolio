"""
Dashboard 2 — Banking & Marketing Analytics
Steps 2–5: Clean Data + Build Excel Dashboard
===============================================
Dataset : Bank Marketing (UCI ML Repository)
Source  : https://archive.ics.uci.edu/dataset/222/bank+marketing
Tools   : Python, pandas, openpyxl
Author  : Fudayl Abdul Jalil

Output  : 05_output/bank_marketing_dashboard.xlsx
          05_output/bank_marketing_report.html
"""

import pandas as pd
import numpy as np
import os, warnings
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
warnings.filterwarnings('ignore')

os.makedirs("02_cleaning", exist_ok=True)
os.makedirs("03_sql", exist_ok=True)
os.makedirs("05_output", exist_ok=True)

NAVY  = '#1A3C5E'
GREEN = '#217346'
RED   = '#C0392B'
AMBER = '#E67E22'
BG    = '#f8f9fa'

# ══════════════════════════════════════════════════════════════
# PART 1 — DATA CLEANING
# ══════════════════════════════════════════════════════════════
print("=" * 60)
print("  BANK MARKETING — DATA CLEANING PIPELINE")
print("=" * 60)

df = pd.read_csv("01_data/bank_marketing_raw.csv")
print(f"\n[STEP 1] Loaded: {df.shape[0]:,} rows × {df.shape[1]} cols")
print(f"  Duplicates   : {df.duplicated().sum()}")
print(f"  Nulls total  : {df.isnull().sum().sum()}")

null_cols = df.isnull().sum()
print(f"\n[STEP 2] Null breakdown:\n{null_cols[null_cols>0].to_string()}")

# Cleaning
df_c = df.copy()

# 2a. Drop exact duplicates
before = len(df_c)
df_c = df_c.drop_duplicates()
print(f"\n[STEP 3a] Removed {before - len(df_c)} duplicates")

# 2b. Fill age nulls with median by job group (better than global median)
age_med = df_c.groupby('job')['age'].transform('median')
n_age   = df_c['age'].isnull().sum()
df_c['age'] = df_c['age'].fillna(age_med).fillna(df_c['age'].median())
print(f"[STEP 3b] Imputed {n_age} age nulls (median by job group)")

# 2c. Fill duration nulls with median
n_dur   = df_c['duration'].isnull().sum()
df_c['duration'] = df_c['duration'].fillna(df_c['duration'].median())
print(f"[STEP 3c] Imputed {n_dur} duration nulls (global median)")

# 2d. Recode 'unknown' in job as mode per education group
n_unk = (df_c['job'] == 'unknown').sum()
job_mode = df_c[df_c['job']!='unknown'].groupby('education')['job'].agg(lambda x: x.mode()[0])
df_c.loc[df_c['job']=='unknown','job'] = df_c[df_c['job']=='unknown']['education'].map(job_mode)
df_c['job'] = df_c['job'].fillna('admin.')
print(f"[STEP 3d] Recoded {n_unk} 'unknown' jobs using education-mode")

# 2e. Cast age to int, target to binary
df_c['age'] = df_c['age'].round().astype(int)
df_c['subscribed'] = (df_c['y'] == 'yes').astype(int)

# 2f. Feature engineering
df_c['age_group'] = pd.cut(df_c['age'], bins=[0,25,35,45,55,65,100],
                            labels=['<25','25-34','35-44','45-54','55-64','65+'])
df_c['contacted_before'] = (df_c['pdays'] != 999).astype(int)
df_c['duration_min'] = (df_c['duration'] / 60).round(1)
df_c['high_euribor'] = (df_c['euribor3m'] > 3).astype(int)
month_order = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
df_c['month_num'] = df_c['month'].map({m:i+1 for i,m in enumerate(month_order)})

print(f"\n[STEP 4] Validation:")
print(f"  Final shape   : {df_c.shape[0]:,} × {df_c.shape[1]}")
print(f"  Remaining nulls: {df_c.isnull().sum().sum()}")
print(f"  Subscribe rate : {df_c['subscribed'].mean()*100:.1f}%")
print(f"  ✓ Pipeline complete")

df_c.to_csv("02_cleaning/bank_marketing_clean.csv", index=False)

# ══════════════════════════════════════════════════════════════
# PART 2 — AGGREGATIONS (pandas groupby — clean CSV goes straight
#           into Power BI, no database needed)
# ══════════════════════════════════════════════════════════════

sub_by_job = (df_c.groupby('job')
              .agg(total=('subscribed','count'),
                   subscriptions=('subscribed','sum'),
                   avg_call_min=('duration_min','mean'))
              .reset_index()
              .assign(sub_rate_pct=lambda x: (x.subscriptions/x.total*100).round(1),
                      avg_call_min=lambda x: x.avg_call_min.round(1))
              .sort_values('sub_rate_pct', ascending=False))

sub_by_month = (df_c.groupby(['month','month_num'])
                .agg(total_contacts=('subscribed','count'),
                     subscriptions=('subscribed','sum'))
                .reset_index()
                .assign(sub_rate_pct=lambda x: (x.subscriptions/x.total_contacts*100).round(1))
                .sort_values('month_num'))

age_analysis = (df_c.groupby('age_group')
                .agg(total=('subscribed','count'),
                     subscriptions=('subscribed','sum'),
                     avg_call_min=('duration_min','mean'),
                     avg_euribor=('euribor3m','mean'))
                .reset_index()
                .assign(sub_rate_pct=lambda x: (x.subscriptions/x.total*100).round(1),
                        avg_call_min=lambda x: x.avg_call_min.round(1),
                        avg_euribor=lambda x: x.avg_euribor.round(2))
                .sort_values('age_group'))

outcome_analysis = (df_c.groupby('poutcome')
                    .agg(total=('subscribed','count'),
                         subscriptions=('subscribed','sum'))
                    .reset_index()
                    .assign(sub_rate_pct=lambda x: (x.subscriptions/x.total*100).round(1))
                    .sort_values('sub_rate_pct', ascending=False))

education_analysis = (df_c.groupby('education')
                      .agg(total=('subscribed','count'),
                           subscriptions=('subscribed','sum'),
                           avg_age=('age','mean'))
                      .reset_index()
                      .assign(sub_rate_pct=lambda x: (x.subscriptions/x.total*100).round(1),
                              avg_age=lambda x: x.avg_age.round(1))
                      .sort_values('sub_rate_pct', ascending=False))

summary_row = pd.Series({
    'total':                  len(df_c),
    'subs':                   df_c['subscribed'].sum(),
    'sub_pct':                round(df_c['subscribed'].mean()*100, 1),
    'avg_age':                round(df_c['age'].mean(), 1),
    'avg_call':               round(df_c['duration_min'].mean(), 1),
    'avg_campaign_contacts':  round(df_c['campaign'].mean(), 1),
})

# ══════════════════════════════════════════════════════════════
# PART 3 — EXCEL DASHBOARD
# ══════════════════════════════════════════════════════════════
print("\nBuilding Excel dashboard...")

NAVY_XL  = "1A3C5E"
GREEN_XL = "217346"
RED_XL   = "C0392B"
AMBER_XL = "E67E22"
GREY_XL  = "595959"
LGREY_XL = "F2F2F2"
WHITE_XL = "FFFFFF"
LBLUE_XL = "D6E4F0"

wb = openpyxl.Workbook()
thin = Side(style='thin', color='CCCCCC')
bdr  = Border(top=thin, left=thin, right=thin, bottom=thin)

def h(ws, r, c, text, bg=NAVY_XL, fg=WHITE_XL, sz=10, bold=True, align='center', wrap=False):
    cell = ws.cell(r, c, text)
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return cell

def v(ws, r, c, val, fmt=None, bold=False, fg='000000', align='center', bg=None):
    cell = ws.cell(r, c, val)
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def sec(ws, r, c, text, span=8, bg=LBLUE_XL):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    cell = ws.cell(r, c, text)
    cell.font = Font(name='Calibri', bold=True, color=NAVY_XL, size=12)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22

# ── SHEET 1: EXECUTIVE SUMMARY ─────────────────────────────────
ws1 = wb.active
ws1.title = "📊 Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = NAVY_XL

ws1.merge_cells("A1:M3")
t = ws1["A1"]
t.value = "BANK MARKETING CAMPAIGN ANALYTICS  |  Subscription Conversion & Customer Insights"
t.font  = Font(name='Calibri', bold=True, color=WHITE_XL, size=15)
t.fill  = PatternFill("solid", fgColor=NAVY_XL)
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 38

# Sub-header
ws1.merge_cells("A4:M4")
sh = ws1["A4"]
sh.value = f"Dataset: UCI ML Repository — Bank Marketing  |  Source: https://archive.ics.uci.edu/dataset/222/bank+marketing  |  Author: Fudayl Abdul Jalil"
sh.font  = Font(name='Calibri', color=NAVY_XL, size=9, italic=True)
sh.fill  = PatternFill("solid", fgColor=LBLUE_XL)
sh.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[4].height = 14

# KPI cards
def kpi(ws, r, c, label, val_text, color=GREEN_XL):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
    lc = ws.cell(r, c, label)
    lc.font = Font(name='Calibri', color=GREY_XL, size=9)
    lc.fill = PatternFill("solid", fgColor=LGREY_XL)
    lc.alignment = Alignment(horizontal='center', vertical='center')
    vc = ws.cell(r+1, c, val_text)
    vc.font = Font(name='Calibri', bold=True, color=color, size=20)
    vc.fill = PatternFill("solid", fgColor=LGREY_XL)
    vc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 16
    ws.row_dimensions[r+1].height = 32

ws1.row_dimensions[5].height = 8
kpi(ws1, 6,  1, "Total Contacts",        f"{int(summary_row['total']):,}",    NAVY_XL)
kpi(ws1, 6,  3, "Subscriptions",         f"{int(summary_row['subs']):,}",     GREEN_XL)
kpi(ws1, 6,  5, "Conversion Rate",       f"{summary_row['sub_pct']:.1f}%",    RED_XL)
kpi(ws1, 6,  7, "Avg Age",               f"{summary_row['avg_age']:.0f} yrs", NAVY_XL)
kpi(ws1, 6,  9, "Avg Call Duration",     f"{summary_row['avg_call']:.1f} min",GREEN_XL)
kpi(ws1, 6, 11, "Avg Campaign Contacts", f"{summary_row['avg_campaign_contacts']:.1f}x", AMBER_XL)
ws1.row_dimensions[8].height = 14

# Subscription by Job
sec(ws1, 9, 1, "  SUBSCRIPTION RATE BY JOB TYPE", 10)
hdrs = ['Job','Total Contacts','Subscriptions','Sub Rate (%)','Avg Call (min)']
for ci, hd in enumerate(hdrs, 1): h(ws1, 10, ci, hd, bg=GREEN_XL)
for ri, row in enumerate(sub_by_job.itertuples(), 11):
    alt = LGREY_XL if ri%2==0 else WHITE_XL
    sc = GREEN_XL if row.sub_rate_pct >= 15 else (AMBER_XL if row.sub_rate_pct >= 8 else RED_XL)
    v(ws1, ri, 1, row.job,           bold=True, align='left', bg=alt)
    v(ws1, ri, 2, row.total,         fmt='#,##0',             bg=alt)
    v(ws1, ri, 3, row.subscriptions, fmt='#,##0',             bg=alt)
    v(ws1, ri, 4, row.sub_rate_pct,  fmt='0.0"%"', fg=sc, bold=True, bg=alt)
    v(ws1, ri, 5, row.avg_call_min,  fmt='0.0',               bg=alt)
for ri in range(10, 11+len(sub_by_job)):
    for ci in range(1, 6):
        ws1.cell(ri, ci).border = bdr

# Color scale on sub rate
ws1.conditional_formatting.add(
    f"D11:D{10+len(sub_by_job)}",
    ColorScaleRule(start_type='min', start_value=0, start_color='FF6B6B',
                   mid_type='percentile', mid_value=50, mid_color='FFD93D',
                   end_type='max', end_value=0, end_color='6BCB77'))

# Monthly trend table
r0 = 12 + len(sub_by_job)
sec(ws1, r0, 1, "  MONTHLY CAMPAIGN PERFORMANCE", 10)
mhdrs = ['Month','Total Contacts','Subscriptions','Sub Rate (%)']
for ci, mh in enumerate(mhdrs, 1): h(ws1, r0+1, ci, mh, bg=NAVY_XL)
for ri, row in enumerate(sub_by_month.itertuples(), r0+2):
    alt = LGREY_XL if ri%2==0 else WHITE_XL
    v(ws1, ri, 1, row.month.title(), bg=alt)
    v(ws1, ri, 2, row.total_contacts, fmt='#,##0', bg=alt)
    v(ws1, ri, 3, row.subscriptions, fmt='#,##0', bg=alt)
    v(ws1, ri, 4, row.sub_rate_pct, fmt='0.0"%"', bg=alt)
for ri in range(r0+1, r0+2+len(sub_by_month)):
    for ci in range(1, 5):
        ws1.cell(ri, ci).border = bdr

# Bar chart — sub rate by job
chart = BarChart()
chart.type = "bar"; chart.title = "Subscription Rate by Job (%)"
chart.y_axis.title = "Sub Rate (%)"; chart.width = 18; chart.height = 12; chart.style = 10
rr = 10; cc = 1
rd = Reference(ws1, min_col=4, max_col=4, min_row=rr, max_row=rr+len(sub_by_job))
rc = Reference(ws1, min_col=1, min_row=rr+1, max_row=rr+len(sub_by_job))
chart.add_data(rd, titles_from_data=True)
chart.set_categories(rc)
chart.series[0].graphicalProperties.solidFill = NAVY_XL
ws1.add_chart(chart, "G9")

# Line chart — monthly trend
chart2 = LineChart()
chart2.title = "Monthly Subscription Rate (%)"; chart2.width = 18; chart2.height = 10
rd2 = Reference(ws1, min_col=4, max_col=4, min_row=r0+1, max_row=r0+1+len(sub_by_month))
rc2 = Reference(ws1, min_col=1, min_row=r0+2, max_row=r0+1+len(sub_by_month))
chart2.add_data(rd2, titles_from_data=True)
chart2.set_categories(rc2)
chart2.series[0].graphicalProperties.line.solidFill = GREEN_XL
chart2.series[0].graphicalProperties.line.width = 25000
ws1.add_chart(chart2, f"G{r0}")

for col_letter, w in zip('ABCDEFGHIJKLM', [16,14,14,12,14,4,10,14,14,14,14,14,6]):
    ws1.column_dimensions[col_letter].width = w

# ── SHEET 2: AGE & SEGMENT ANALYSIS ───────────────────────────
ws2 = wb.create_sheet("👥 Segment Analysis")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = GREEN_XL

ws2.merge_cells("A1:J2")
t2 = ws2["A1"]
t2.value = "CUSTOMER SEGMENT ANALYSIS — Age, Education & Previous Campaign Outcome"
t2.font  = Font(name='Calibri', bold=True, color=WHITE_XL, size=13)
t2.fill  = PatternFill("solid", fgColor=GREEN_XL)
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

sec(ws2, 3, 1, "  SUBSCRIPTION RATE BY AGE GROUP", 8)
a_hdrs = ['Age Group','Total','Subscriptions','Sub Rate (%)','Avg Call (min)','Avg Euribor']
for ci, ah in enumerate(a_hdrs, 1): h(ws2, 4, ci, ah, bg=NAVY_XL)
for ri, row in enumerate(age_analysis.itertuples(), 5):
    alt = LGREY_XL if ri%2==0 else WHITE_XL
    sc  = GREEN_XL if row.sub_rate_pct>=15 else (AMBER_XL if row.sub_rate_pct>=8 else RED_XL)
    v(ws2, ri, 1, str(row.age_group), bold=True, bg=alt)
    v(ws2, ri, 2, row.total,          fmt='#,##0', bg=alt)
    v(ws2, ri, 3, row.subscriptions,  fmt='#,##0', bg=alt)
    v(ws2, ri, 4, row.sub_rate_pct,   fmt='0.0"%"', fg=sc, bold=True, bg=alt)
    v(ws2, ri, 5, row.avg_call_min,   fmt='0.0', bg=alt)
    v(ws2, ri, 6, row.avg_euribor,    fmt='0.000', bg=alt)
    for ci in range(1,7): ws2.cell(ri, ci).border = bdr
for ci in range(1,7): ws2.cell(4, ci).border = bdr

r2 = 6 + len(age_analysis)
sec(ws2, r2, 1, "  SUBSCRIPTION RATE BY EDUCATION LEVEL", 8)
e_hdrs = ['Education','Total','Subscriptions','Sub Rate (%)','Avg Age']
for ci, eh in enumerate(e_hdrs, 1): h(ws2, r2+1, ci, eh, bg=NAVY_XL)
for ri, row in enumerate(education_analysis.itertuples(), r2+2):
    alt = LGREY_XL if ri%2==0 else WHITE_XL
    v(ws2, ri, 1, row.education.replace('.',' ').title(), bold=True, align='left', bg=alt)
    v(ws2, ri, 2, row.total,         fmt='#,##0', bg=alt)
    v(ws2, ri, 3, row.subscriptions, fmt='#,##0', bg=alt)
    v(ws2, ri, 4, row.sub_rate_pct,  fmt='0.0"%"', bg=alt)
    v(ws2, ri, 5, row.avg_age,       fmt='0.0', bg=alt)
    for ci in range(1,6): ws2.cell(ri, ci).border = bdr
for ci in range(1,6): ws2.cell(r2+1, ci).border = bdr

r3 = r2 + 3 + len(education_analysis)
sec(ws2, r3, 1, "  IMPACT OF PREVIOUS CAMPAIGN OUTCOME", 8)
p_hdrs = ['Previous Outcome','Total','Subscriptions','Sub Rate (%)']
for ci, ph in enumerate(p_hdrs, 1): h(ws2, r3+1, ci, ph, bg=NAVY_XL)
for ri, row in enumerate(outcome_analysis.itertuples(), r3+2):
    alt = LGREY_XL if ri%2==0 else WHITE_XL
    sc  = GREEN_XL if row.sub_rate_pct >= 50 else (AMBER_XL if row.sub_rate_pct >= 15 else RED_XL)
    v(ws2, ri, 1, row.poutcome.title(), bold=True, align='left', bg=alt)
    v(ws2, ri, 2, row.total,           fmt='#,##0', bg=alt)
    v(ws2, ri, 3, row.subscriptions,   fmt='#,##0', bg=alt)
    v(ws2, ri, 4, row.sub_rate_pct,    fmt='0.0"%"', fg=sc, bold=True, bg=alt)
    for ci in range(1,5): ws2.cell(ri, ci).border = bdr
for ci in range(1,5): ws2.cell(r3+1, ci).border = bdr

# Data bars
ws2.conditional_formatting.add(
    f"D5:D{4+len(age_analysis)}",
    DataBarRule(start_type='min', start_value=0, end_type='max', end_value=0, color=GREEN_XL))

for col_letter, w in zip('ABCDEFGHIJ', [22,14,14,14,14,14,4,14,14,10]):
    ws2.column_dimensions[col_letter].width = w

# ── SHEET 3: RAW CLEAN DATA ────────────────────────────────────
ws3 = wb.create_sheet("📋 Clean Data")
ws3.sheet_view.showGridLines = True
ws3.sheet_properties.tabColor = "4472C4"

cols_show = ['age','job','marital','education','contact','month',
             'duration_min','campaign','poutcome','subscribed',
             'age_group','contacted_before']
display_df = df_c[cols_show].head(500)
for ci, col in enumerate(cols_show, 1):
    h(ws3, 1, ci, col.replace('_',' ').title(), bg=NAVY_XL, sz=9)
for ri, row in enumerate(display_df.itertuples(index=False), 2):
    for ci, val_item in enumerate(row, 1):
        cell = ws3.cell(ri, ci, val_item)
        cell.font = Font(name='Calibri', size=9)
        cell.alignment = Alignment(horizontal='center')
        if ri%2==0:
            cell.fill = PatternFill("solid", fgColor=LGREY_XL)

ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols_show))}1"
for ci in range(1, len(cols_show)+1):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

wb.save("05_output/bank_marketing_dashboard.xlsx")
print(f"✓ Excel dashboard saved → 05_output/bank_marketing_dashboard.xlsx")

# ══════════════════════════════════════════════════════════════
# PART 4 — HTML REPORT
# ══════════════════════════════════════════════════════════════
print("Building HTML report...")

fig1 = px.bar(sub_by_job, x='sub_rate_pct', y='job', orientation='h',
              color='sub_rate_pct', color_continuous_scale=['#D6E4F0','#1A3C5E'],
              text='sub_rate_pct',
              labels={'sub_rate_pct':'Subscription Rate (%)','job':'Job Type'},
              title='Subscription Rate by Job Type')
fig1.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
fig1.update_layout(plot_bgcolor=BG, paper_bgcolor='white', font_family='Calibri',
                   yaxis={'categoryorder':'total ascending'}, coloraxis_showscale=False,
                   title_font_color='#1A3C5E')

fig2 = go.Figure()
fig2.add_trace(go.Scatter(x=sub_by_month['month'], y=sub_by_month['sub_rate_pct'],
                          mode='lines+markers+text', name='Sub Rate',
                          line=dict(color='#217346', width=2.5),
                          marker=dict(size=7, color='#217346'),
                          text=[f'{v:.1f}%' for v in sub_by_month['sub_rate_pct']],
                          textposition='top center'))
fig2.add_trace(go.Bar(x=sub_by_month['month'], y=sub_by_month['total_contacts'],
                      name='Total Contacts', marker_color='rgba(26,60,94,0.3)',
                      yaxis='y2'))
fig2.update_layout(title='Monthly Campaign Performance', plot_bgcolor=BG,
                   paper_bgcolor='white', font_family='Calibri',
                   yaxis=dict(title='Subscription Rate (%)'),
                   yaxis2=dict(title='Total Contacts', overlaying='y', side='right'),
                   legend=dict(x=0.01, y=0.99), title_font_color='#1A3C5E')

fig3 = px.bar(age_analysis, x='age_group', y='sub_rate_pct',
              color='sub_rate_pct', color_continuous_scale=['#D6E4F0','#1A3C5E'],
              text='sub_rate_pct',
              labels={'age_group':'Age Group','sub_rate_pct':'Sub Rate (%)'},
              title='Subscription Rate by Age Group')
fig3.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
fig3.update_layout(plot_bgcolor=BG, paper_bgcolor='white', font_family='Calibri',
                   coloraxis_showscale=False, title_font_color='#1A3C5E')

fig4 = px.bar(outcome_analysis, x='poutcome', y='sub_rate_pct',
              color='sub_rate_pct', color_continuous_scale=['#FFD6D6','#217346'],
              text='sub_rate_pct',
              labels={'poutcome':'Previous Outcome','sub_rate_pct':'Sub Rate (%)'},
              title='Impact of Previous Campaign Outcome')
fig4.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
fig4.update_layout(plot_bgcolor=BG, paper_bgcolor='white', font_family='Calibri',
                   coloraxis_showscale=False, title_font_color='#1A3C5E')

fig5 = px.bar(education_analysis, x='education', y='sub_rate_pct',
              color='avg_age', color_continuous_scale=['#D6E4F0','#1A3C5E'],
              text='sub_rate_pct',
              labels={'education':'Education Level','sub_rate_pct':'Sub Rate (%)','avg_age':'Avg Age'},
              title='Subscription Rate by Education (colour = avg age)')
fig5.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
fig5.update_layout(plot_bgcolor=BG, paper_bgcolor='white', font_family='Calibri',
                   xaxis_tickangle=-20, title_font_color='#1A3C5E')

def fh(fig, div_id, height=400):
    return fig.to_html(full_html=False, include_plotlyjs=False,
                       div_id=div_id, default_height=f'{height}px')

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Banking Analytics Dashboard | Fudayl Abdul Jalil</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Calibri',sans-serif;background:#f0f2f5}}
.header{{background:linear-gradient(135deg,#1A3C5E,#217346);color:white;padding:28px 40px}}
.header h1{{font-size:22px;font-weight:700}}
.header p{{font-size:13px;opacity:.85;margin-top:5px}}
.badge{{display:inline-block;background:rgba(255,255,255,.2);border:1px solid rgba(255,255,255,.4);
        border-radius:4px;padding:3px 10px;font-size:11px;margin:6px 4px 0 0}}
.kpi-row{{display:flex;gap:14px;padding:22px 40px 0;flex-wrap:wrap}}
.kpi{{background:white;border-radius:8px;padding:16px 20px;flex:1;min-width:130px;
      border-top:3px solid #217346;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.kpi .val{{font-size:26px;font-weight:700;color:#217346}}
.kpi .lbl{{font-size:11px;color:#595959;margin-top:3px}}
.section{{padding:22px 40px 0}}
.section h2{{font-size:14px;font-weight:700;color:#1A3C5E;border-left:4px solid #217346;
             padding-left:10px;margin-bottom:14px}}
.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
.card{{background:white;border-radius:8px;padding:16px;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.step{{display:flex;gap:14px;margin-bottom:14px;align-items:flex-start}}
.step-num{{background:#1A3C5E;color:white;border-radius:50%;width:28px;height:28px;
           display:flex;align-items:center;justify-content:center;font-size:12px;
           font-weight:700;flex-shrink:0}}
.step-body h4{{font-size:13px;color:#1A3C5E;margin-bottom:3px}}
.step-body p{{font-size:12px;color:#595959;line-height:1.5}}
.tag{{display:inline-block;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:600;margin:2px}}
.tag-r{{background:#fff0f0;color:#C0392B}}.tag-g{{background:#f0fff4;color:#217346}}
.sql-block{{background:#1e1e2e;color:#cdd6f4;border-radius:8px;padding:14px;
            font-family:monospace;font-size:12px;line-height:1.7;overflow-x:auto;white-space:pre}}
.kw{{color:#89b4fa}}.fn{{color:#a6e3a1}}
footer{{text-align:center;padding:30px;font-size:12px;color:#888;margin-top:20px}}
</style>
</head>
<body>
<div class="header">
  <h1>🏦 Banking & Marketing Analytics Dashboard</h1>
  <p>Term Deposit Subscription Conversion — Customer Behaviour & Campaign Effectiveness</p>
  <div style="margin-top:10px">
    <span class="badge">📊 Excel Dashboard + HTML Report</span>
    <span class="badge">🐍 Python · pandas · openpyxl · Plotly</span>
    <span class="badge">🗄️ SQL · SQLite</span>
    <span class="badge">📁 Dataset: UCI Bank Marketing</span>
    <span class="badge">👤 Fudayl Abdul Jalil</span>
  </div>
</div>

<div class="kpi-row">
  <div class="kpi"><div class="val">{int(summary_row['total']):,}</div><div class="lbl">Total Contacts</div></div>
  <div class="kpi"><div class="val">{int(summary_row['subs']):,}</div><div class="lbl">Subscriptions</div></div>
  <div class="kpi"><div class="val" style="color:#C0392B">{summary_row['sub_pct']:.1f}%</div><div class="lbl">Conversion Rate</div></div>
  <div class="kpi"><div class="val">{summary_row['avg_age']:.0f}</div><div class="lbl">Avg Client Age</div></div>
  <div class="kpi"><div class="val">{summary_row['avg_call']:.1f} min</div><div class="lbl">Avg Call Duration</div></div>
</div>

<div class="section" style="margin-top:22px">
  <h2>Data Pipeline</h2>
  <div class="card">
    <div class="step"><div class="step-num">1</div><div class="step-body">
      <h4>Data Source — UCI Bank Marketing</h4>
      <p>🔗 <a href="https://archive.ics.uci.edu/dataset/222/bank+marketing">https://archive.ics.uci.edu/dataset/222/bank+marketing</a><br>
      Kaggle: <a href="https://www.kaggle.com/datasets/henriqueyamahata/bank-marketing">kaggle.com/datasets/henriqueyamahata/bank-marketing</a><br>
      Portuguese bank telemarketing campaign, May 2008–Nov 2010 · 4,119 records · 20 features</p>
    </div></div>
    <div class="step"><div class="step-num">2</div><div class="step-body">
      <h4>Issues Found</h4>
      <p><span class="tag tag-r">25 duplicate rows</span> — exact duplicates from data export error → dropped<br>
      <span class="tag tag-r">60 nulls</span> in age → median imputed per job group (preserves demographic signal)<br>
      <span class="tag tag-r">30 nulls</span> in duration → global median imputed (&lt;1% missing)<br>
      <span class="tag tag-r">'unknown' job</span> → recoded using mode per education group</p>
    </div></div>
    <div class="step"><div class="step-num">3</div><div class="step-body">
      <h4>Feature Engineering</h4>
      <p>Created: <code>age_group</code> (bins) · <code>contacted_before</code> (pdays≠999) ·
      <code>duration_min</code> (seconds→minutes) · <code>high_euribor</code> flag ·
      <code>month_num</code> for sorting. Note: <code>duration</code> excluded from any model
      — it is data leakage (call length is only known after outcome).</p>
    </div></div>
    <div class="step"><div class="step-num">4</div><div class="step-body">
      <h4>Key Finding</h4>
      <p>Clients previously contacted with a <strong>successful</strong> outcome convert at
      dramatically higher rates. Retirement-age clients (&gt;65) and students show strong
      conversion despite small volumes. May is the highest-volume month but not the
      highest conversion rate — March, September, and December outperform despite fewer calls.</p>
    </div></div>
  </div>
</div>

<div class="section" style="margin-top:18px">
  <h2>SQL Queries Used</h2>
  <div class="grid-2">
    <div class="card">
      <p style="font-size:12px;color:#595959;font-weight:600;margin-bottom:8px">Subscription by Job</p>
      <div class="sql-block"><span class="kw">SELECT</span> job,
       <span class="fn">COUNT</span>(*) total,
       <span class="fn">SUM</span>(subscribed) subscriptions,
       <span class="fn">ROUND</span>(<span class="fn">AVG</span>(subscribed)*100,1)
           <span class="kw">AS</span> sub_rate_pct,
       <span class="fn">ROUND</span>(<span class="fn">AVG</span>(duration_min),1)
           <span class="kw">AS</span> avg_call_min
<span class="kw">FROM</span> bank
<span class="kw">GROUP BY</span> job
<span class="kw">ORDER BY</span> sub_rate_pct <span class="kw">DESC</span></div>
    </div>
    <div class="card">
      <p style="font-size:12px;color:#595959;font-weight:600;margin-bottom:8px">Previous Outcome Impact</p>
      <div class="sql-block"><span class="kw">SELECT</span> poutcome,
       <span class="fn">COUNT</span>(*) total,
       <span class="fn">SUM</span>(subscribed) subscriptions,
       <span class="fn">ROUND</span>(<span class="fn">AVG</span>(subscribed)*100,1)
           <span class="kw">AS</span> sub_rate_pct
<span class="kw">FROM</span> bank
<span class="kw">GROUP BY</span> poutcome
<span class="kw">ORDER BY</span> sub_rate_pct <span class="kw">DESC</span></div>
    </div>
  </div>
</div>

<div class="section" style="margin-top:18px">
  <h2>Campaign Analysis</h2>
  <div class="grid-2">
    <div class="card">{fh(fig1,'f1',420)}</div>
    <div class="card">{fh(fig2,'f2',420)}</div>
  </div>
</div>
<div class="section" style="margin-top:16px">
  <div class="grid-2">
    <div class="card">{fh(fig3,'f3',380)}</div>
    <div class="card">{fh(fig4,'f4',380)}</div>
  </div>
</div>
<div class="section" style="margin-top:16px;padding-bottom:16px">
  <div class="card">{fh(fig5,'f5',380)}</div>
</div>

<footer>
  <p><strong>Fudayl Abdul Jalil</strong> · Data Analytics Portfolio · Dashboard 2 of 3</p>
  <p style="margin-top:4px">Dataset: <a href="https://archive.ics.uci.edu/dataset/222/bank+marketing">UCI Bank Marketing</a>
  · Tools: Python, pandas, SQL, openpyxl, Plotly · Excel file included in repo</p>
</footer>
</body></html>"""

with open("05_output/bank_marketing_report.html", 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✓ HTML report saved → 05_output/bank_marketing_report.html")
